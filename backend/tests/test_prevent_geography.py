from __future__ import annotations

import csv
import io
import unittest
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import ValidationError

from app.schemas.prevent_record import PreventRecordCreate, PreventRecordListFilters
from app.services.prevent_engine import compute_prevent_10y
from app.services.prevent_records import (
    PREVENT_EXPORT_HEADERS,
    create_prevent_record,
    export_prevent_records_csv,
    export_prevent_records_xlsx,
)


BASE_PAYLOAD = {
    "age": 55,
    "sex": "male",
    "total_cholesterol": 220,
    "hdl": 45,
    "sbp": 132,
    "egfr": 92,
    "bmi": 28,
    "diabetes": True,
    "smoker": False,
    "antihypertensive_use": False,
    "statin_use": False,
    "physician_name": "Dr. Example",
    "physician_specialty": "Cardiologia",
}

GEOGRAPHY_PAYLOAD = {
    "patient_province_code": "17",
    "patient_province_name": "Pichincha",
    "patient_canton_code": "1701",
    "patient_canton_name": "Quito",
    "patient_area_type": "urban",
    "patient_geo_source": "self_reported",
}

GEOGRAPHY_HEADERS = [
    "patient_province_code",
    "patient_province_name",
    "patient_canton_code",
    "patient_canton_name",
    "patient_area_type",
    "patient_geo_source",
]


class FakeSession:
    def __init__(self) -> None:
        self.record = None

    def add(self, record) -> None:
        self.record = record

    def commit(self) -> None:
        return None

    def refresh(self, record) -> None:
        if record.id is None:
            record.id = uuid4()


class EmptyQuery:
    def filter(self, *args, **kwargs):
        return self

    def order_by(self, *args, **kwargs):
        return self

    def all(self):
        return []


class EmptyExportSession:
    def query(self, *args, **kwargs):
        return EmptyQuery()


class PreventGeographyTest(unittest.TestCase):
    def test_create_record_with_complete_geography(self) -> None:
        db = FakeSession()
        payload = PreventRecordCreate(**BASE_PAYLOAD, **GEOGRAPHY_PAYLOAD)

        response = create_prevent_record(db=db, payload=payload)

        self.assertIsNotNone(db.record)
        self.assertEqual(db.record.patient_province, "Pichincha")
        self.assertEqual(db.record.patient_province_code, "17")
        self.assertEqual(db.record.patient_province_name, "Pichincha")
        self.assertEqual(db.record.patient_canton_code, "1701")
        self.assertEqual(db.record.patient_canton_name, "Quito")
        self.assertEqual(db.record.patient_area_type, "urban")
        self.assertEqual(db.record.patient_geo_source, "self_reported")
        self.assertEqual(response.patient_province_code, "17")
        self.assertEqual(response.patient_canton_code, "1701")
        self.assertIsNotNone(response.cvd_risk)
        self.assertIsNotNone(response.cvd_risk_30y)

    def test_create_record_without_geography(self) -> None:
        db = FakeSession()
        payload = PreventRecordCreate(**BASE_PAYLOAD)

        response = create_prevent_record(db=db, payload=payload)

        self.assertIsNotNone(db.record)
        self.assertIsNone(db.record.patient_province_code)
        self.assertIsNone(db.record.patient_province_name)
        self.assertIsNone(db.record.patient_canton_code)
        self.assertIsNone(db.record.patient_canton_name)
        self.assertIsNone(db.record.patient_area_type)
        self.assertIsNone(db.record.patient_geo_source)
        self.assertIsNone(response.patient_province_code)
        self.assertIsNone(response.patient_geo_source)

    def test_invalid_area_type_is_rejected(self) -> None:
        with self.assertRaises(ValidationError):
            PreventRecordCreate(**BASE_PAYLOAD, patient_area_type="suburban")

    def test_invalid_geo_source_is_rejected(self) -> None:
        with self.assertRaises(ValidationError):
            PreventRecordCreate(**BASE_PAYLOAD, patient_geo_source="manual")

    def test_canton_code_requires_province_code_and_canton_name(self) -> None:
        with self.assertRaises(ValidationError):
            PreventRecordCreate(
                **BASE_PAYLOAD,
                patient_canton_code="1701",
                patient_canton_name="Quito",
            )
        with self.assertRaises(ValidationError):
            PreventRecordCreate(
                **BASE_PAYLOAD,
                patient_province_code="17",
                patient_province_name="Pichincha",
                patient_canton_code="1701",
            )

    def test_province_code_requires_province_name(self) -> None:
        with self.assertRaises(ValidationError):
            PreventRecordCreate(**BASE_PAYLOAD, patient_province_code="17")

    def test_csv_export_includes_geography_columns(self) -> None:
        self.assertEqual(PREVENT_EXPORT_HEADERS[-6:], GEOGRAPHY_HEADERS)

        csv_content = export_prevent_records_csv(
            db=EmptyExportSession(),
            filters=PreventRecordListFilters(),
        )
        header = next(csv.reader(io.StringIO(csv_content.lstrip("\ufeff")), delimiter=";"))

        self.assertEqual(header[-6:], GEOGRAPHY_HEADERS)

    def test_xlsx_export_includes_geography_columns(self) -> None:
        xlsx_content = export_prevent_records_xlsx(
            db=EmptyExportSession(),
            filters=PreventRecordListFilters(),
        )
        workbook = load_workbook(io.BytesIO(xlsx_content), read_only=True)
        worksheet = workbook.active
        header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

        self.assertEqual(header[-6:], GEOGRAPHY_HEADERS)

    def test_prevent_risks_do_not_change_with_geography_fields(self) -> None:
        clinical_payload = {
            key: value
            for key, value in BASE_PAYLOAD.items()
            if key not in {"physician_name", "physician_specialty"}
        }
        _, base_result, base_warnings = compute_prevent_10y(clinical_payload)
        _, geo_result, geo_warnings = compute_prevent_10y(
            {**clinical_payload, **GEOGRAPHY_PAYLOAD},
        )

        self.assertEqual(base_warnings, geo_warnings)
        for key in ("cvd_10y", "ascvd_10y", "hf_10y", "cvd_30y", "ascvd_30y", "hf_30y"):
            self.assertEqual(base_result[key], geo_result[key])


if __name__ == "__main__":
    unittest.main()
