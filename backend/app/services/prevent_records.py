import csv
import io
import logging
from datetime import datetime, timezone
from uuid import UUID, uuid4

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.prevent_record import PreventRecord
from app.schemas.prevent_record import (
    PreventRecordCreate,
    PreventRecordCreateResponse,
    PreventRecordDetailResponse,
    PreventRecordListFilters,
    PreventRecordListItem,
    PreventRecordListResponse,
)
from app.services.auth_users import AuthenticatedUser
from app.services.clinical_recommendations import (
    build_clinical_interpretation,
    normalize_clinical_interpretation,
)
from app.services.prevent_engine import (
    calculate_prevent_age,
    classify_risk,
    compute_prevent_10y,
)
from app.services.prevent_validation import invalid_outcomes_from_warnings, validate_prevent_inputs
from app.services.prevent_coefficients import PREVENT_ENGINE_STATUS, PREVENT_METHOD_NOTE

logger = logging.getLogger(__name__)

PREVENT_EXPORT_HEADERS = [
    "id",
    "fecha",
    "edad",
    "sexo",
    "medico",
    "especialidad_medica",
    "colesterol_total",
    "hdl",
    "presion_sistolica",
    "egfr",
    "imc",
    "diabetes",
    "tabaquismo",
    "uso_estatinas",
    "uso_antihipertensivos",
    "uacr",
    "hba1c",
    "sdi",
    "riesgo_cvd_10y",
    "riesgo_ascvd_10y",
    "riesgo_ic_10y",
    "riesgo_cvd_30y",
    "riesgo_ascvd_30y",
    "riesgo_ic_30y",
    "edad_cardiovascular_equivalente",
    "variante_modelo",
    "dominios_recomendacion",
    "base_lipidos",
    "guia_lipidos",
    "base_presion_arterial",
    "guia_presion_arterial",
    "base_insuficiencia_cardiaca",
    "guia_insuficiencia_cardiaca",
    "base_renal_cardiorrenal",
    "guia_renal_cardiorrenal",
    "patient_province_code",
    "patient_province_name",
    "patient_canton_code",
    "patient_canton_name",
    "patient_area_type",
    "patient_geo_source",
    "cobertura_sanitaria",
    "nivel_educativo",
    "situacion_laboral",
    "etnia",
    "nivel_socioeconomico",
]


def _safe_payload_for_log(payload: dict[str, object]) -> dict[str, object]:
    return {
        key: value
        for key, value in payload.items()
        if key not in {"physician_name", "physician_specialty", "notes"}
    }


def _payload_value(record: PreventRecord, key: str):
    payload = record.input_payload_json or {}
    return payload.get(key)


def _result_payload_value(record: PreventRecord, key: str):
    payload = record.input_payload_json or {}
    results = payload.get("results") or {}
    return results.get(key)


def _clinical_interpretation_from_record(record: PreventRecord) -> dict[str, object] | None:
    payload = record.input_payload_json or {}
    stored_interpretation = payload.get("clinical_interpretation")
    if isinstance(stored_interpretation, dict):
        normalized = normalize_clinical_interpretation(stored_interpretation)
        if normalized is not None:
            return normalized

    extracted = _extract_record_results(record)
    engine_input = {
        **payload,
        "age": record.patient_age,
        "sex": record.patient_sex,
        "total_cholesterol": record.total_cholesterol,
        "hdl": record.hdl_cholesterol,
        "sbp": record.systolic_bp,
        "egfr": record.egfr,
        "bmi": record.bmi,
        "diabetes": record.diabetes,
        "smoker": record.smoker,
        "statin_use": record.statin_use,
        "antihypertensive_use": record.antihypertensive_use,
    }
    warnings = payload.get("warnings") if isinstance(payload.get("warnings"), list) else None
    return build_clinical_interpretation(
        prevent_result={
            "cvd_risk": extracted["cvd_risk"],
            "ascvd_risk": extracted["ascvd_risk"],
            "hf_risk": extracted["hf_risk"],
            "prevent_age": extracted["prevent_age"],
            "model_variant": extracted["model_variant"],
        },
        input_payload=engine_input,
        warnings=warnings,
    )


def _domain_traceability_for_export(record: PreventRecord) -> dict[str, str | None]:
    clinical = _clinical_interpretation_from_record(record) or {}
    domains = clinical.get("domain_recommendations") or []
    traceability: dict[str, dict[str, object]] = {}
    if isinstance(domains, list):
        for domain in domains:
            if isinstance(domain, dict):
                key = str(domain.get("key") or "")
                traceability[key] = domain

    def value(key: str, field: str) -> str | None:
        domain = traceability.get(key)
        if not domain:
            return None
        raw_value = domain.get(field)
        return str(raw_value) if raw_value is not None else None

    domain_types = [
        str(domain.get("domain_type") or domain.get("key"))
        for domain in traceability.values()
        if isinstance(domain, dict)
    ]
    return {
        "dominios_recomendacion": "|".join(domain_types) if domain_types else None,
        "base_lipidos": value("lipids", "recommendation_basis"),
        "guia_lipidos": value("lipids", "guideline_context"),
        "base_presion_arterial": value("blood_pressure", "recommendation_basis"),
        "guia_presion_arterial": value("blood_pressure", "guideline_context"),
        "base_insuficiencia_cardiaca": value("heart_failure", "recommendation_basis"),
        "guia_insuficiencia_cardiaca": value("heart_failure", "guideline_context"),
        "base_renal_cardiorrenal": value("renal", "recommendation_basis"),
        "guia_renal_cardiorrenal": value("renal", "guideline_context"),
    }


def _column_or_payload(record: PreventRecord, column_value, payload_key: str):
    return column_value if column_value is not None else _payload_value(record, payload_key)


def _format_exact_risk_for_csv(value) -> float | None:
    if value is None:
        return None
    return float(value)


def _format_decimal_for_regional_csv(value) -> str | None:
    if value is None:
        return None
    return f"{float(value):.15g}".replace(".", ",")


def _format_boolean_for_csv(value: bool | None) -> str | None:
    if value is None:
        return None
    return "Sí" if value else "No"


def _format_value_for_regional_csv(value):
    if isinstance(value, bool):
        return _format_boolean_for_csv(value)
    if isinstance(value, float):
        return _format_decimal_for_regional_csv(value)
    return value


def _build_prevent_export_row(record: PreventRecord) -> list[object]:
    extracted = _extract_record_results(record)
    uacr = _column_or_payload(record, record.uacr, "uacr")
    hba1c = _column_or_payload(record, record.hba1c, "hba1c")
    sdi = _column_or_payload(record, record.sdi, "sdi")
    model_variant = record.model_variant_used or _payload_value(record, "model_variant")
    cvd_risk = record.cvd_risk_10y
    if cvd_risk is None and record.risk_10y is not None:
        cvd_risk = record.risk_10y * 100
    if cvd_risk is None:
        cvd_risk = _result_payload_value(record, "cvd_10y")
    ascvd_risk = (
        record.ascvd_risk_10y
        if record.ascvd_risk_10y is not None
        else _result_payload_value(record, "ascvd_10y")
    )
    hf_risk = (
        record.hf_risk_10y
        if record.hf_risk_10y is not None
        else _result_payload_value(record, "hf_10y")
    )
    cvd_risk_30y = (
        record.cvd_risk_30y
        if record.cvd_risk_30y is not None
        else _result_payload_value(record, "cvd_30y")
    )
    ascvd_risk_30y = (
        record.ascvd_risk_30y
        if record.ascvd_risk_30y is not None
        else _result_payload_value(record, "ascvd_30y")
    )
    hf_risk_30y = (
        record.hf_risk_30y
        if record.hf_risk_30y is not None
        else _result_payload_value(record, "hf_30y")
    )
    prevent_age = (
        record.prevent_age
        if record.prevent_age is not None
        else _result_payload_value(record, "prevent_age")
    )
    traceability = _domain_traceability_for_export(record)

    return [
        str(record.id),
        record.created_at.isoformat(),
        record.patient_age,
        record.patient_sex,
        record.physician_name,
        record.physician_specialty,
        record.total_cholesterol,
        record.hdl_cholesterol,
        record.systolic_bp,
        record.egfr,
        record.bmi,
        record.diabetes,
        record.smoker,
        record.statin_use,
        record.antihypertensive_use,
        uacr,
        hba1c,
        sdi,
        _format_exact_risk_for_csv(
            cvd_risk if cvd_risk is not None else extracted["cvd_risk"],
        ),
        _format_exact_risk_for_csv(
            ascvd_risk if ascvd_risk is not None else extracted["ascvd_risk"],
        ),
        _format_exact_risk_for_csv(
            hf_risk if hf_risk is not None else extracted["hf_risk"],
        ),
        _format_exact_risk_for_csv(
            cvd_risk_30y if cvd_risk_30y is not None else extracted["cvd_30y"],
        ),
        _format_exact_risk_for_csv(
            ascvd_risk_30y if ascvd_risk_30y is not None else extracted["ascvd_30y"],
        ),
        _format_exact_risk_for_csv(
            hf_risk_30y if hf_risk_30y is not None else extracted["hf_30y"],
        ),
        float(prevent_age) if prevent_age is not None else None,
        model_variant or extracted["model_variant"],
        traceability["dominios_recomendacion"],
        traceability["base_lipidos"],
        traceability["guia_lipidos"],
        traceability["base_presion_arterial"],
        traceability["guia_presion_arterial"],
        traceability["base_insuficiencia_cardiaca"],
        traceability["guia_insuficiencia_cardiaca"],
        traceability["base_renal_cardiorrenal"],
        traceability["guia_renal_cardiorrenal"],
        record.patient_province_code,
        record.patient_province_name,
        record.patient_canton_code,
        record.patient_canton_name,
        record.patient_area_type,
        record.patient_geo_source,
        record.patient_health_coverage,
        record.patient_education_level,
        record.patient_employment_status,
        record.patient_ethnicity,
        record.patient_socioeconomic_level,
    ]


def _extract_record_results(record: PreventRecord) -> dict[str, float | str | None]:
    payload = record.input_payload_json or {}
    results = payload.get("results") or {}
    cvd_risk = record.cvd_risk_10y
    if cvd_risk is None and record.risk_10y is not None:
        cvd_risk = record.risk_10y * 100
    if cvd_risk is None:
        cvd_risk = results.get("cvd_10y")

    return {
        "cvd_risk": cvd_risk,
        "ascvd_risk": record.ascvd_risk_10y
        if record.ascvd_risk_10y is not None
        else results.get("ascvd_10y"),
        "hf_risk": record.hf_risk_10y
        if record.hf_risk_10y is not None
        else results.get("hf_10y"),
        "cvd_30y": record.cvd_risk_30y
        if record.cvd_risk_30y is not None
        else results.get("cvd_30y"),
        "ascvd_30y": record.ascvd_risk_30y
        if record.ascvd_risk_30y is not None
        else results.get("ascvd_30y"),
        "hf_30y": record.hf_risk_30y
        if record.hf_risk_30y is not None
        else results.get("hf_30y"),
        "prevent_age": record.prevent_age
        if record.prevent_age is not None
        else results.get("prevent_age"),
        "model_variant": record.model_variant_used or payload.get("model_variant"),
    }


def _build_prevent_records_query(
    db: Session,
    filters: PreventRecordListFilters,
    *,
    apply_status_filter: bool = True,
):
    query = db.query(PreventRecord)

    if filters.owner_doctor_id is not None:
        query = query.filter(PreventRecord.owner_doctor_id == filters.owner_doctor_id)
    if filters.visibility_scope is not None:
        query = query.filter(PreventRecord.visibility_scope == filters.visibility_scope)
    if filters.visibility_scopes:
        query = query.filter(PreventRecord.visibility_scope.in_(filters.visibility_scopes))
    if filters.date_from is not None:
        query = query.filter(func.date(PreventRecord.created_at) >= filters.date_from)
    if filters.date_to is not None:
        query = query.filter(func.date(PreventRecord.created_at) <= filters.date_to)
    if filters.physician_name:
        query = query.filter(
            PreventRecord.physician_name.ilike(f"%{filters.physician_name.strip()}%"),
        )
    if filters.diabetes is not None:
        query = query.filter(PreventRecord.diabetes == filters.diabetes)
    if filters.smoker is not None:
        query = query.filter(PreventRecord.smoker == filters.smoker)
    if filters.patient_province_code:
        query = query.filter(PreventRecord.patient_province_code == filters.patient_province_code)
    if filters.patient_canton_code:
        query = query.filter(PreventRecord.patient_canton_code == filters.patient_canton_code)
    if filters.patient_area_type:
        query = query.filter(PreventRecord.patient_area_type == filters.patient_area_type)
    if filters.patient_geo_source:
        query = query.filter(PreventRecord.patient_geo_source == filters.patient_geo_source)
    if filters.patient_health_coverage:
        query = query.filter(PreventRecord.patient_health_coverage == filters.patient_health_coverage)
    if filters.patient_education_level:
        query = query.filter(PreventRecord.patient_education_level == filters.patient_education_level)
    if filters.patient_employment_status:
        query = query.filter(PreventRecord.patient_employment_status == filters.patient_employment_status)
    if filters.patient_ethnicity:
        query = query.filter(PreventRecord.patient_ethnicity == filters.patient_ethnicity)
    if filters.patient_socioeconomic_level:
        query = query.filter(
            PreventRecord.patient_socioeconomic_level == filters.patient_socioeconomic_level,
        )
    if filters.model_variant is not None:
        query = query.filter(
            PreventRecord.input_payload_json["model_variant"].astext == filters.model_variant,
        )
    if apply_status_filter:
        if filters.record_status == "active":
            query = query.filter(PreventRecord.is_deleted.is_(False))
        elif filters.record_status == "archived":
            query = query.filter(PreventRecord.is_deleted.is_(True))

    return query


def _record_matches_access_scope(
    record: PreventRecord,
    *,
    owner_doctor_id: UUID | None = None,
    visibility_scope: str | None = None,
    visibility_scopes: list[str] | None = None,
) -> bool:
    if owner_doctor_id is not None and record.owner_doctor_id != owner_doctor_id:
        return False
    if visibility_scope is not None and record.visibility_scope != visibility_scope:
        return False
    if visibility_scopes is not None and record.visibility_scope not in visibility_scopes:
        return False
    return True


def _validate_requested_variant(engine_input: dict[str, object]) -> str | None:
    requested_variant = engine_input.get("model_variant")
    if requested_variant is None:
        return None

    if requested_variant == "uacr" and engine_input.get("uacr") is None:
        raise ValueError("El modelo UACR requiere un valor de UACR.")
    if requested_variant == "hba1c" and engine_input.get("hba1c") is None:
        raise ValueError("El modelo HbA1c requiere un valor de HbA1c.")
    if requested_variant == "sdi" and engine_input.get("sdi") is None:
        raise ValueError("El modelo SDI requiere un valor de SDI.")

    return str(requested_variant)


def _build_engine_warnings(warnings: list[str] | None) -> list[dict[str, object]]:
    return [
        {
            "field": "engine",
            "field_label": "Motor PREVENT",
            "value": None,
            "unit": None,
            "min": None,
            "max": None,
            "range": None,
            "outcomes": ["cvd", "ascvd", "hf"],
            "message": str(warning),
            "severity": "warning",
        }
        for warning in warnings or []
    ]


def _evaluate_prevent_payload(
    payload: PreventRecordCreate,
    *,
    debug: bool = False,
) -> dict[str, object]:
    engine_input = payload.model_dump(by_alias=True, exclude_none=True)
    requested_variant = _validate_requested_variant(engine_input)
    logger.info("PREVENT input payload received: %s", _safe_payload_for_log(engine_input))
    model_variant, result, warnings = compute_prevent_10y(engine_input, requested_variant)
    validation_warnings = validate_prevent_inputs(engine_input)
    invalid_outcomes = invalid_outcomes_from_warnings(validation_warnings)
    if "cvd" in invalid_outcomes:
        result["cvd_10y"] = None
        result["cvd_30y"] = None
    if "ascvd" in invalid_outcomes:
        result["ascvd_10y"] = None
        result["ascvd_30y"] = None
    if "hf" in invalid_outcomes:
        result["hf_10y"] = None
        result["hf_30y"] = None
    combined_warnings = [
        *_build_engine_warnings(warnings),
        *validation_warnings,
    ] or None
    logger.info(
        "PREVENT raw engine result: %s",
        {
            "model_variant": model_variant,
            "result": result,
            "warnings": combined_warnings,
        },
    )
    cvd_risk = result["cvd_10y"]
    ascvd_risk = result["ascvd_10y"]
    hf_risk = result["hf_10y"]
    cvd_risk_30y = result["cvd_30y"]
    ascvd_risk_30y = result["ascvd_30y"]
    hf_risk_30y = result["hf_30y"]
    prevent_age = calculate_prevent_age(cvd_risk, engine_input.get("sex"))
    cvd_category = classify_risk(cvd_risk) if cvd_risk is not None else None
    ascvd_category = classify_risk(ascvd_risk) if ascvd_risk is not None else None
    hf_category = classify_risk(hf_risk) if hf_risk is not None else None
    clinical_interpretation = build_clinical_interpretation(
        prevent_result={
            "cvd_risk": cvd_risk,
            "ascvd_risk": ascvd_risk,
            "hf_risk": hf_risk,
            "cvd_risk_30y": cvd_risk_30y,
            "ascvd_risk_30y": ascvd_risk_30y,
            "hf_risk_30y": hf_risk_30y,
            "prevent_age": prevent_age,
            "model_variant": model_variant,
        },
        input_payload=engine_input,
        warnings=combined_warnings,
    )
    risk_10y_stored = (cvd_risk / 100.0) if cvd_risk is not None else None
    evaluation_debug = None
    if debug:
        evaluation_debug = {
            "cvd_input_summary": {
                "model": f"prevent_{model_variant}_10y",
                "outcome": "cvd",
                "inputs": engine_input,
            },
            "ascvd_input_summary": {
                "model": f"prevent_{model_variant}_10y",
                "outcome": "ascvd",
                "inputs": engine_input,
            },
            "hf_input_summary": {
                "model": f"prevent_{model_variant}_10y",
                "outcome": "hf",
                "inputs": engine_input,
                "requires_bmi": True,
            },
            "model_variant": model_variant,
            "cvd_risk": cvd_risk,
            "ascvd_risk": ascvd_risk,
            "hf_risk": hf_risk,
            "prevent_age": prevent_age,
            "warnings": combined_warnings,
            "clinical_interpretation": clinical_interpretation,
        }

    return {
        "engine_input": engine_input,
        "result": result,
        "model_variant": model_variant,
        "combined_warnings": combined_warnings,
        "cvd_risk": cvd_risk,
        "ascvd_risk": ascvd_risk,
        "hf_risk": hf_risk,
        "cvd_risk_30y": cvd_risk_30y,
        "ascvd_risk_30y": ascvd_risk_30y,
        "hf_risk_30y": hf_risk_30y,
        "prevent_age": prevent_age,
        "cvd_category": cvd_category,
        "ascvd_category": ascvd_category,
        "hf_category": hf_category,
        "clinical_interpretation": clinical_interpretation,
        "risk_10y_stored": risk_10y_stored,
        "evaluation_debug": evaluation_debug,
    }


def calculate_prevent_record_preview(
    payload: PreventRecordCreate,
    *,
    debug: bool = False,
) -> PreventRecordCreateResponse:
    evaluation = _evaluate_prevent_payload(payload, debug=debug)
    return PreventRecordCreateResponse(
        id=uuid4(),
        cvd_risk=evaluation["cvd_risk"],
        ascvd_risk=evaluation["ascvd_risk"],
        hf_risk=evaluation["hf_risk"],
        cvd_risk_30y=evaluation["cvd_risk_30y"],
        ascvd_risk_30y=evaluation["ascvd_risk_30y"],
        hf_risk_30y=evaluation["hf_risk_30y"],
        prevent_age=evaluation["prevent_age"],
        cvd_category=evaluation["cvd_category"],
        ascvd_category=evaluation["ascvd_category"],
        hf_category=evaluation["hf_category"],
        model_variant=str(evaluation["model_variant"]),
        risk_10y=evaluation["risk_10y_stored"],
        risk_category=evaluation["cvd_category"],
        engine_status=PREVENT_ENGINE_STATUS,
        methodology_note=PREVENT_METHOD_NOTE,
        warnings=evaluation["combined_warnings"],
        clinical_interpretation=evaluation["clinical_interpretation"],
        debug=evaluation["evaluation_debug"],
        message="Prevent risk calculated without saving",
        patient_province_code=payload.patient_province_code,
        patient_province_name=payload.patient_province_name,
        patient_canton_code=payload.patient_canton_code,
        patient_canton_name=payload.patient_canton_name,
        patient_area_type=payload.patient_area_type,
        patient_geo_source=payload.patient_geo_source,
        patient_health_coverage=payload.patient_health_coverage,
        patient_education_level=payload.patient_education_level,
        patient_employment_status=payload.patient_employment_status,
        patient_ethnicity=payload.patient_ethnicity,
        patient_socioeconomic_level=payload.patient_socioeconomic_level,
    )


def create_prevent_record(
    db: Session,
    payload: PreventRecordCreate,
    *,
    debug: bool = False,
    current_user: AuthenticatedUser | None = None,
) -> PreventRecordCreateResponse:
    request_id = str(uuid4())
    evaluation = _evaluate_prevent_payload(payload, debug=debug)
    engine_input = evaluation["engine_input"]
    result = evaluation["result"]
    model_variant = evaluation["model_variant"]
    combined_warnings = evaluation["combined_warnings"]
    cvd_risk = evaluation["cvd_risk"]
    ascvd_risk = evaluation["ascvd_risk"]
    hf_risk = evaluation["hf_risk"]
    cvd_risk_30y = evaluation["cvd_risk_30y"]
    ascvd_risk_30y = evaluation["ascvd_risk_30y"]
    hf_risk_30y = evaluation["hf_risk_30y"]
    prevent_age = evaluation["prevent_age"]
    cvd_category = evaluation["cvd_category"]
    ascvd_category = evaluation["ascvd_category"]
    hf_category = evaluation["hf_category"]
    clinical_interpretation = evaluation["clinical_interpretation"]
    risk_10y_stored = evaluation["risk_10y_stored"]
    evaluation_debug = evaluation["evaluation_debug"]

    record_data = payload.model_dump(
        exclude_none=True,
        exclude={
            "risk_10y",
            "model_variant",
            "zip_code",
            "input_payload_json",
        },
    )
    if current_user is not None:
        record_traceability = {
            "created_by_user_id": current_user.user.id,
            "owner_doctor_id": current_user.doctor_profile.id
            if current_user.doctor_profile is not None
            else None,
            "source_type": "doctor",
            "user_type": "doctor",
            "visibility_scope": "doctor_private",
            "created_modality": "doctor_calculator",
            "request_id": request_id,
        }
    else:
        record_traceability = {
            "source_type": "public",
            "user_type": "anonymous",
            "visibility_scope": "public_anonymous",
            "created_modality": "public_calculator",
            "request_id": request_id,
        }

    record = PreventRecord(
        **record_data,
        **record_traceability,
        risk_10y=risk_10y_stored,
        risk_category=cvd_category,
        model_variant_used=model_variant,
        cvd_risk_10y=cvd_risk,
        ascvd_risk_10y=ascvd_risk,
        hf_risk_10y=hf_risk,
        cvd_risk_30y=cvd_risk_30y,
        ascvd_risk_30y=ascvd_risk_30y,
        hf_risk_30y=hf_risk_30y,
        cvd_category=cvd_category,
        ascvd_category=ascvd_category,
        hf_category=hf_category,
        prevent_age=prevent_age,
        input_payload_json={
            **engine_input,
            "model_variant": model_variant,
            "warnings": combined_warnings,
            "clinical_interpretation": clinical_interpretation,
            "results": {
                **result,
                "prevent_age": prevent_age,
            },
        },
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    logger.info(
        "PREVENT outcome results generated: %s",
        {
            "cvd_risk": cvd_risk,
            "ascvd_risk": ascvd_risk,
            "hf_risk": hf_risk,
            "cvd_risk_30y": cvd_risk_30y,
            "ascvd_risk_30y": ascvd_risk_30y,
            "hf_risk_30y": hf_risk_30y,
            "prevent_age": prevent_age,
            "model_variant": model_variant,
            "warnings": combined_warnings,
            "clinical_interpretation": clinical_interpretation,
        },
    )
    response_payload = {
        "id": record.id,
        "cvd_risk": cvd_risk,
        "ascvd_risk": ascvd_risk,
        "hf_risk": hf_risk,
        "cvd_risk_30y": cvd_risk_30y,
        "ascvd_risk_30y": ascvd_risk_30y,
        "hf_risk_30y": hf_risk_30y,
        "prevent_age": prevent_age,
        "cvd_category": cvd_category,
        "ascvd_category": ascvd_category,
        "hf_category": hf_category,
        "model_variant": model_variant,
        "risk_10y": record.risk_10y,
        "risk_category": record.risk_category,
        "engine_status": PREVENT_ENGINE_STATUS,
        "methodology_note": PREVENT_METHOD_NOTE,
        "warnings": combined_warnings,
        "clinical_interpretation": clinical_interpretation,
        "debug": evaluation_debug,
        "message": "Prevent risk calculated successfully",
        "patient_province_code": record.patient_province_code,
        "patient_province_name": record.patient_province_name,
        "patient_canton_code": record.patient_canton_code,
        "patient_canton_name": record.patient_canton_name,
        "patient_area_type": record.patient_area_type,
        "patient_geo_source": record.patient_geo_source,
        "patient_health_coverage": record.patient_health_coverage,
        "patient_education_level": record.patient_education_level,
        "patient_employment_status": record.patient_employment_status,
        "patient_ethnicity": record.patient_ethnicity,
        "patient_socioeconomic_level": record.patient_socioeconomic_level,
        "source_type": record.source_type,
        "user_type": record.user_type,
        "visibility_scope": record.visibility_scope,
        "created_modality": record.created_modality,
        "request_id": record.request_id,
        "created_by_user_id": record.created_by_user_id,
        "owner_doctor_id": record.owner_doctor_id,
        "patient_id": record.patient_id,
        "public_session_id": record.public_session_id,
    }
    logger.info("PREVENT response payload: %s", response_payload)
    return PreventRecordCreateResponse(
        id=record.id,
        cvd_risk=cvd_risk,
        ascvd_risk=ascvd_risk,
        hf_risk=hf_risk,
        cvd_risk_30y=cvd_risk_30y,
        ascvd_risk_30y=ascvd_risk_30y,
        hf_risk_30y=hf_risk_30y,
        prevent_age=prevent_age,
        cvd_category=cvd_category,
        ascvd_category=ascvd_category,
        hf_category=hf_category,
        model_variant=model_variant,
        risk_10y=record.risk_10y,
        risk_category=record.risk_category,
        engine_status=PREVENT_ENGINE_STATUS,
        methodology_note=PREVENT_METHOD_NOTE,
        warnings=combined_warnings,
        clinical_interpretation=clinical_interpretation,
        debug=evaluation_debug,
        message="Prevent risk calculated successfully",
        patient_province_code=record.patient_province_code,
        patient_province_name=record.patient_province_name,
        patient_canton_code=record.patient_canton_code,
        patient_canton_name=record.patient_canton_name,
        patient_area_type=record.patient_area_type,
        patient_geo_source=record.patient_geo_source,
        patient_health_coverage=record.patient_health_coverage,
        patient_education_level=record.patient_education_level,
        patient_employment_status=record.patient_employment_status,
        patient_ethnicity=record.patient_ethnicity,
        patient_socioeconomic_level=record.patient_socioeconomic_level,
        source_type=record.source_type,
        user_type=record.user_type,
        visibility_scope=record.visibility_scope,
        created_modality=record.created_modality,
        request_id=record.request_id,
        created_by_user_id=record.created_by_user_id,
        owner_doctor_id=record.owner_doctor_id,
        patient_id=record.patient_id,
        public_session_id=record.public_session_id,
    )


def get_prevent_record(db: Session, record_id: UUID) -> PreventRecord | None:
    return db.get(PreventRecord, record_id)


def archive_prevent_record(db: Session, record_id: UUID) -> PreventRecord | None:
    record = get_prevent_record(db, record_id)
    if record is None:
        return None

    if not record.is_deleted:
        record.is_deleted = True
        record.deleted_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(record)
    return record


def restore_prevent_record(db: Session, record_id: UUID) -> PreventRecord | None:
    record = get_prevent_record(db, record_id)
    if record is None:
        return None

    if record.is_deleted:
        record.is_deleted = False
        record.deleted_at = None
        db.commit()
        db.refresh(record)
    return record


def list_prevent_records(
    db: Session,
    filters: PreventRecordListFilters,
) -> PreventRecordListResponse:
    page = max(filters.page, 1)
    page_size = min(max(filters.page_size, 1), 100)
    query = _build_prevent_records_query(db, filters)
    all_status_query = _build_prevent_records_query(
        db,
        filters,
        apply_status_filter=False,
    )
    total = query.count()
    active_total = all_status_query.filter(PreventRecord.is_deleted.is_(False)).count()
    archived_total = all_status_query.filter(PreventRecord.is_deleted.is_(True)).count()
    records = (
        query.order_by(PreventRecord.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items: list[PreventRecordListItem] = []
    for record in records:
        extracted = _extract_record_results(record)
        items.append(
            PreventRecordListItem(
                id=record.id,
                created_at=record.created_at,
                is_deleted=record.is_deleted,
                deleted_at=record.deleted_at,
                patient_age=record.patient_age,
                patient_sex=record.patient_sex,
                physician_name=record.physician_name,
                diabetes=record.diabetes,
                smoker=record.smoker,
                patient_province_code=record.patient_province_code,
                patient_province_name=record.patient_province_name,
                patient_canton_code=record.patient_canton_code,
                patient_canton_name=record.patient_canton_name,
                patient_area_type=record.patient_area_type,
                patient_geo_source=record.patient_geo_source,
                patient_health_coverage=record.patient_health_coverage,
                patient_education_level=record.patient_education_level,
                patient_employment_status=record.patient_employment_status,
                patient_ethnicity=record.patient_ethnicity,
                patient_socioeconomic_level=record.patient_socioeconomic_level,
                cvd_risk=extracted["cvd_risk"],
                ascvd_risk=extracted["ascvd_risk"],
                hf_risk=extracted["hf_risk"],
                cvd_risk_30y=extracted["cvd_30y"],
                ascvd_risk_30y=extracted["ascvd_30y"],
                hf_risk_30y=extracted["hf_30y"],
                model_variant=extracted["model_variant"],
                created_by_user_id=record.created_by_user_id,
                owner_doctor_id=record.owner_doctor_id,
                patient_id=record.patient_id,
                public_session_id=record.public_session_id,
                source_type=record.source_type,
                user_type=record.user_type,
                visibility_scope=record.visibility_scope,
                created_modality=record.created_modality,
                request_id=record.request_id,
            ),
        )

    return PreventRecordListResponse(
        items=items,
        total=total,
        active_total=active_total,
        archived_total=archived_total,
        page=page,
        page_size=page_size,
    )


def get_prevent_record_detail(
    db: Session,
    record_id: UUID,
    *,
    owner_doctor_id: UUID | None = None,
    visibility_scope: str | None = None,
    visibility_scopes: list[str] | None = None,
) -> PreventRecordDetailResponse | None:
    record = get_prevent_record(db, record_id)
    if record is None:
        return None
    if not _record_matches_access_scope(
        record,
        owner_doctor_id=owner_doctor_id,
        visibility_scope=visibility_scope,
        visibility_scopes=visibility_scopes,
    ):
        return None

    extracted = _extract_record_results(record)
    return PreventRecordDetailResponse(
        id=record.id,
        created_at=record.created_at,
        updated_at=record.updated_at,
        is_deleted=record.is_deleted,
        deleted_at=record.deleted_at,
        patient_age=record.patient_age,
        patient_sex=record.patient_sex,
        patient_country=record.patient_country,
        patient_province=record.patient_province,
        patient_province_code=record.patient_province_code,
        patient_province_name=record.patient_province_name,
        patient_canton_code=record.patient_canton_code,
        patient_canton_name=record.patient_canton_name,
        patient_area_type=record.patient_area_type,
        patient_geo_source=record.patient_geo_source,
        patient_health_coverage=record.patient_health_coverage,
        patient_education_level=record.patient_education_level,
        patient_employment_status=record.patient_employment_status,
        patient_ethnicity=record.patient_ethnicity,
        patient_socioeconomic_level=record.patient_socioeconomic_level,
        total_cholesterol=record.total_cholesterol,
        hdl_cholesterol=record.hdl_cholesterol,
        ldl_cholesterol=record.ldl_cholesterol,
        systolic_bp=record.systolic_bp,
        diabetes=record.diabetes,
        smoker=record.smoker,
        bmi=record.bmi,
        egfr=record.egfr,
        statin_use=record.statin_use,
        antihypertensive_use=record.antihypertensive_use,
        physician_name=record.physician_name,
        physician_specialty=record.physician_specialty,
        physician_city=record.physician_city,
        risk_10y=record.risk_10y,
        risk_category=record.risk_category,
        model_variant_used=record.model_variant_used,
        cvd_risk_10y=record.cvd_risk_10y,
        ascvd_risk_10y=record.ascvd_risk_10y,
        hf_risk_10y=record.hf_risk_10y,
        cvd_risk_30y=extracted["cvd_30y"],
        ascvd_risk_30y=extracted["ascvd_30y"],
        hf_risk_30y=extracted["hf_30y"],
        cvd_category=record.cvd_category,
        ascvd_category=record.ascvd_category,
        hf_category=record.hf_category,
        prevent_age=record.prevent_age,
        engine_version=record.engine_version,
        source_org=record.source_org,
        initiative_name=record.initiative_name,
        director_name=record.director_name,
        consent_for_research=record.consent_for_research,
        notes=record.notes,
        input_payload_json=record.input_payload_json,
        cvd_risk=extracted["cvd_risk"],
        ascvd_risk=extracted["ascvd_risk"],
        hf_risk=extracted["hf_risk"],
        cvd_30y=extracted["cvd_30y"],
        ascvd_30y=extracted["ascvd_30y"],
        hf_30y=extracted["hf_30y"],
        model_variant=extracted["model_variant"],
        clinical_interpretation=_clinical_interpretation_from_record(record),
        created_by_user_id=record.created_by_user_id,
        owner_doctor_id=record.owner_doctor_id,
        patient_id=record.patient_id,
        public_session_id=record.public_session_id,
        source_type=record.source_type,
        user_type=record.user_type,
        visibility_scope=record.visibility_scope,
        created_modality=record.created_modality,
        request_id=record.request_id,
    )


def export_prevent_records_csv(
    db: Session,
    filters: PreventRecordListFilters,
) -> str:
    query = _build_prevent_records_query(db, filters)
    records = query.order_by(PreventRecord.created_at.desc()).all()

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    writer.writerow(PREVENT_EXPORT_HEADERS)

    for record in records:
        writer.writerow(
            [
                _format_value_for_regional_csv(value)
                for value in _build_prevent_export_row(record)
            ],
        )

    return "\ufeff" + buffer.getvalue()


def export_prevent_records_xlsx(
    db: Session,
    filters: PreventRecordListFilters,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    query = _build_prevent_records_query(db, filters)
    records = query.order_by(PreventRecord.created_at.desc()).all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PREVENT"
    worksheet.append(PREVENT_EXPORT_HEADERS)

    for record in records:
        worksheet.append(_build_prevent_export_row(record))

    header_fill = PatternFill("solid", fgColor="E5F3FB")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="1C6D9B")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    risk_headers = {
        "riesgo_cvd_10y",
        "riesgo_ascvd_10y",
        "riesgo_ic_10y",
        "riesgo_cvd_30y",
        "riesgo_ascvd_30y",
        "riesgo_ic_30y",
    }
    decimal_headers = {
        "colesterol_total",
        "hdl",
        "presion_sistolica",
        "egfr",
        "imc",
        "uacr",
        "hba1c",
        "sdi",
        "edad_cardiovascular_equivalente",
        *risk_headers,
    }
    for column_index, header in enumerate(PREVENT_EXPORT_HEADERS, start=1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = max(len(header) + 2, 14)
        if header in decimal_headers:
            number_format = "0.000000" if header in risk_headers else "0.######"
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column_index).number_format = number_format

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
